from django.conf import settings
from django.db import IntegrityError

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import date, datetime, timedelta
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail

from mailing_machine.models import *

import requests
import json
import os

# Create your views here.
# Settings
API_KEY = "Goes here" #TODO: Add your API Key here.


def establish_wobo_connection(model='team', identifier='', params={}, api_key=None):
    # Here we are establishing a HTTP GET connection to WorkBoard's REST API. To keep this dynamic we allow two
    # variables: model, which is the kind of data model we want to pull and identifier, which is to specify certain
    # items we might want to pull from the API.
    # The 'Authorization' header is necessary as we need to prove our identity to WorkBoard.
    url = 'https://www.myworkboard.com/wb/apis/{}/{}'.format(model, identifier)
    print('Connecting to {}. Please wait as we generate the output.'.format(url))
    
    if api_key:
        headers = {'Authorization': "bearer {}".format(api_key)}
    else:
        # Detecon internal use with Bradley's API key.
        headers = {'Authorization': "bearer {}".format(BRADLEY_WOBO)}
    
    req = requests.get(url=url,
                       headers=headers,
                       params=params,
                       )
    
    data = json.loads(req.text)
    
    return {'status_code': req.status_code, 'data': data}


def get_objectives(identifier=None, api_key=None):
    '''
    This function retrieves all Objectives from the WoBo API. We send an HTTP request and retrieve all Objectives
    associated with the organization.
    :param identifier: Might be used as a selector to retrieve the data for a certain team
    :param api_key: Necessary to authenticate against WoBo. Also used to specify the organization.
    :return: Returns a dict with all data coming from the API.
    '''
    if identifier:
        response = establish_wobo_connection(model='goal', identifier=identifier, api_key=api_key)
    else:
        response = establish_wobo_connection(model='goal', api_key=api_key)
    
    if response['status_code'] == 200:
        data = response['data']
        
        return data
    else:
        raise BrokenPipeError('There was an error processing your Objectives request.')


def get_teams(identifier='', api_key=None):
    '''
    Returns all teams that live within an organization.
    :param identifier: Can be used to retrive data for a certain team.
    :param api_key: Necessary to authenticate against WoBo. Also used to specify the organization.
    :return: Returns a dict with all data coming from the API.
    '''
    if identifier:
        response = establish_wobo_connection(model='team', identifier=identifier, api_key=api_key)
    else:
        response = establish_wobo_connection(model='team', api_key=api_key)
    
    if response['status_code'] == 200:
        teams = response['data']['data']['team']
        data = []

        for team in teams:
    
            try:
                t = WoBoTeam.objects.get(pk=team['team_id'])
                data.append(t)
    
            except WoBoTeam.DoesNotExist:
        
                try:
                    try:
                        company = WoBoCompany.objects.get(api_key=api_key)
                    except WoBoCompany.DoesNotExist:
                        company = None
                   
                    t = WoBoTeam.objects.create(
                        id=team['team_id'],
                        company=company,
                        manager=team['team_owner'],
                        parent_team_id=team['parent_team_id'],
                        team_name=team['team_name']
                    )
                    data.append(t)
                    
        
                except IntegrityError:
                    pass

        return data
    
    else:
        raise BrokenPipeError('There was an error processing your Teams request.')
    
    
def get_user_teams(api_key=None, identifier=''):
    '''
    This fuction is used to get all teams a user is part of. However, this only works as a one to many relationship and
    implies that one has to loop through all users in order to find out their team associations...
    :param api_key: Necessary to authenticate against WoBo. Also used to specify the organization.
    :return: Returns a dict with all data coming from the API.
    '''
    
    if not identifier:
        identifier = establish_wobo_connection(model='user', api_key=api_key)['data']['data']['user']
    
    response = establish_wobo_connection(model='user', identifier='{}/team/'.format(identifier), api_key=api_key)
    
    if response['status_code'] == 200:
        return response['data']
    else:
        print(response)
        return None
    

def get_key_results(identifier='', api_key=None):
    if identifier:
        response = establish_wobo_connection(model='metric', identifier=identifier, api_key=api_key)
    else:
        response = establish_wobo_connection(model='metric', api_key=api_key)
    
    if response['status_code'] == 200:
        data = response['data']
        
        return data
    else:
        raise BrokenPipeError('There was an error processing your Metrics request.')
    

def get_users(identifier='', api_key=None):
    
    params = {'include': 'team_members'}
    data = list()
    
    me = establish_wobo_connection(model='user', api_key=api_key)
    
    if me['status_code'] == 200:
        print(me)
    else:
        return {'status': me['status_code'], 'message': me['message']}
    
    if identifier:
        response = establish_wobo_connection(model='user', identifier=identifier, api_key=api_key, params=params)
    else:
        response = establish_wobo_connection(model='user', api_key=api_key, params=params)
    
    if response['status_code'] == 200:
        users = response['data']['data']['user']
        
        for user in users:
            try:
                u = WoBoTeamMember.objects.get(pk=user['user_id'])
            except WoBoTeamMember.DoesNotExist:
                u = WoBoTeamMember.objects.create(
                    id=user['user_id'],
                    first_name=user['first_name'],
                    last_name=user['last_name'],
                    email=user['email'],
                )
                
                try:
                    team = WoBoTeam.objects.get(pk=user['team_id']).members.add(u)
                except WoBoTeam.DoesNotExist:
                    pass
                
            data.append(u)
        
        return data
        
    else:
        raise BrokenPipeError('There was an error processing your Metrics request.')
    
    
def build_org_chart(identifier='', api_key=None):
    
    users = get_users(identifier, api_key)
    teams = get_teams(identifier, api_key)
    
    for user in users:
        t = get_user_teams(api_key=api_key, identifier=user.pk)
        
        if t is not None:
            print(t.keys())
            for team in t['data']['user']['team']:
                try:
                    WoBoTeam.objects.get(pk=team['team_id']).members.add(user)
                    print("User '{}' was added to team '{}'".format(user.email, team['team_name']))
                except WoBoTeam.DoesNotExist:
                    pass
    
    return {'users': users, 'teams': teams}
    
    
def build_data_sheet(team=None, api_key=None):
    '''
    This function unites all partial functions above so that we can fetch all relevant data from WorkBoard's API and
    process it here.
    :param team: The identifier may be used to specify a certain team and their respective stats.
    :return:
    '''
    
    teams = get_teams(identifier=team, api_key=api_key)
    owners = get_objectives(api_key=api_key)['data']
    print('Connection successful. All data has been fetched. We are now preparing your export.')
    
    wb = Workbook()
    
    # Create an overview of all teams the user is part of.
    ts = wb.active
    ts.title = 'Teams'
    
    col_names = [*teams[0].__dict__.keys()]
    ts.append(col_names)
    for team in teams:
        team = team.__dict__
        team['_state'] = None
        ts.append([*team.values()])
    
    # Create an overview of all objectives the user contributes to.
    obs = wb.create_sheet('Objectives')
    col_names = ['Owner ID', 'Owner', 'Objective ID', 'Objective', 'Date created', 'Date modified', 'Start date',
                 'End date', 'Progress', 'Follow up necessary']
    obs.append(col_names)
    
    # Create an overview of all key results the user contributes to.
    ks = wb.create_sheet('Key Results')
    col_names = ['Owner ID', 'Owner', 'Objective ID', 'Objective', 'Metric name', 'Progress', 'Date created',
                 'Date modified',
                 'Last update', 'Next update', 'Follow up necessary']
    ks.append(col_names)
    
    for owner in owners['goal']:
        for goal in owner['people_goals']:
            li = [owner['user_id'], owner['user_email']]
            li += [goal['goal_id'], goal['goal_name'],
                   datetime.fromtimestamp(int(goal['goal_create_at'])),
                   datetime.fromtimestamp(int(goal['goal_modified_at'])),
                   datetime.fromtimestamp(int(goal['goal_start_date'])),
                   datetime.fromtimestamp(int(goal['goal_target_date'])),
                   float(goal['goal_progress']), ]
            
            # Determine update-overdue status
            if datetime.now() - timedelta(weeks=2) > datetime.fromtimestamp(int(goal['goal_modified_at'])) and \
                    float(goal['goal_progress']) < 100:
                li += [1]
            else:
                li += [0]
            
            obs.append(li)
            
            # Evaluate all key results associated with the objective.
            for metric in goal['goal_metrics']:
                kl = [owner['user_id'], owner['user_email'], goal['goal_id'], goal['goal_name'], metric['metric_name'],
                      float(metric['metric_progress']),
                      datetime.fromtimestamp(int(metric['metric_create_at'])),
                      datetime.fromtimestamp(int(metric['metric_modified_at'])),
                      datetime.fromtimestamp(int(metric['metric_last_update'])),
                      datetime.fromtimestamp(int(metric['metric_next_update']))]
                
                if datetime.now() > datetime.fromtimestamp(int(metric['metric_next_update'])) and float(
                        metric['metric_progress']) < 100:
                    kl += [1]
                else:
                    kl += [0]
                
                ks.append(kl)
    
    # Save the Excel workbook.
    if api_key:
        path = 'media/export/{}.xlsx'.format(api_key)
    else:
        path = 'media/export/wobo_export.xlsx'
    
    wb.save(path)
    
    return path


if __name__ == '__main__':
    build_org_chart(api_key=API_KEY)

